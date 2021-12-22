from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score, confusion_matrix, classification_report, ConfusionMatrixDisplay
from sklearn.exceptions import FitFailedWarning
from scipy.ndimage.interpolation import shift
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import time, os

MNIST_TRAIN_DATA_FILE = "fashion-mnist_train.csv"
MNIST_TEST_DATA_FILE = "fashion-mnist_test.csv"
EXCEL_SHEET_NAME = "Fashion Classifier.xlsx"
MNIST_CF_FOLDER = "Fashion Confusion Matrix Images/"
MNIST_CR_FOLDER = "Fashion Classification Report/"
SHEET_COL_NAMES = [
    "Date",
    "Classifier and Ensembles",
    "Augmented",
    "Accuracy(%)",
    "F1-Score(%)",
    "Precision Score(%)",
    "Recall Score(%)",
    "Training Time (s)",
    "Parameters",
    "Confusion Matrix"
]
MNIST_LABEL_NAMES = [
    "T-shirt/top", 
    "Trouser", 
    "Pullover", 
    "Dress", 
    "Coat", 
    "Sandal", 
    "Shirt", 
    "Sneaker", 
    "Bag", 
    "Ankle boot"
]
IMAGE_WIDTH  = 65.28
IMAGE_HEIGHT = 21.12

# This class calculates and prints out the metrics of a classifier 
# based on the MNIST/FASHION Dataset
class Classifier:
    # init method or constructor 
    def __init__(self):
        self.X_train, self.y_train, self.X_test, self.y_test = self.loadFullMNISTDataCVS()
        
    # Method that loads the mnist data set
    def loadFullMNISTDataCVS(self):
        dftrain = pd.read_csv(MNIST_TRAIN_DATA_FILE)
        dftest = pd.read_csv(MNIST_TEST_DATA_FILE)

        y_train = dftrain['label']
        X_train = dftrain.drop('label', axis = 1)
        X_train = np.array(X_train)
                            
        y_test = dftest['label']
        X_test = dftest.drop('label', axis = 1)
        X_test = np.array(X_test)

        return X_train, y_train, X_test, y_test
    
    # Method to shift and augment the image by given dimension
    def shiftImage(self, image, dx, dy):
        image = image.reshape((28, 28))
        shifted_image = shift(image, [dy, dx], cval=0, mode="constant")
        return shifted_image.reshape([-1])
    
    # Method that gets the exact date and time
    def dateAndTime(self):
        from datetime import datetime
        
        # datetime object containing current date and time
        now = datetime.now()
        
        # dd/mm/YY H:M:S
        self.date = now.strftime("%d/%m/%Y %H:%M:%S")

    # Method that gets parameters that have been changed
    def getParams(self):
        from sklearn import set_config
        # Only print the changed params
        set_config(print_changed_only=True)
        
        # Remove the name and keep only params
        params = str(self.model)
        params = params.replace(self.name, '')
        
        # If there are no params changed
        if not params or params == "()":
            params = "None"

        return params

    # Method that calculates the metrics of the classifier
    def calculateClassifierMetrics(self):
        # Start the counter
        start = time.time()
        
        # Fit the model
        self.model.fit(self.X_train, self.y_train)
        
        # Stop the counter
        stop = time.time()
        
        # Calculate stats of the classifier_model
        self.time = stop - start
        self.y_pred = self.model.predict(self.X_test)
        self.acc = accuracy_score(self.y_test, self.y_pred) * 100
        self.f1 = f1_score(self.y_test, self.y_pred, average='macro') * 100
        self.prec = precision_score(self.y_test, self.y_pred, average='macro') * 100
        self.recall = recall_score(self.y_test, self.y_pred, average='macro') * 100
        self.report = classification_report(self.y_test, self.y_pred, digits=5, output_dict=True, target_names=MNIST_LABEL_NAMES)
        self.c_matrix = ConfusionMatrixDisplay(confusion_matrix=confusion_matrix(self.y_test, self.y_pred), display_labels=MNIST_LABEL_NAMES)
    
    # Method that prints the classifiers information to an excel sheet
    def printToSheet(self, isAugmented):
        # Check if there is already a sheet, if not make one
        if not os.path.isfile(EXCEL_SHEET_NAME):
            book = Workbook()
        else:
            book = load_workbook(EXCEL_SHEET_NAME)
            
        # reference to the active sheet    
        sheet = book.active
        
        # If the spreadsheet was just created, make the headers
        if sheet.max_column == 1:
            i = 1
            for name in SHEET_COL_NAMES:
                sheet.cell(row=1, column=i).value = name
                i += 1
                
        # Insert the data
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1).value = self.date
        sheet.cell(row=row, column=2).value = self.name
        sheet.cell(row=row, column=3).value = isAugmented
        sheet.cell(row=row, column=4).value = float("%0.3f" % (self.acc))
        sheet.cell(row=row, column=5).value = float("%0.3f" % (self.f1))
        sheet.cell(row=row, column=6).value = float("%0.3f" % (self.prec))
        sheet.cell(row=row, column=7).value = float("%0.3f" % (self.recall))
        sheet.cell(row=row, column=8).value = float("%0.1f" % (self.time))
        sheet.cell(row=row, column=9).value = self.getParams()
        sheet.cell(row=row, column=10).value = ' '
        # Insert picture of confusion Matrix
        image = self.convertCMatrixToImage(isAugmented)
        image.anchor = sheet.cell(row=row, column=10).coordinate
        sheet.add_image(image)
        
        book.save(EXCEL_SHEET_NAME)
 
    # Method that converts the classifier's confusion matrix to an image
    def convertCMatrixToImage(self, isAugmented):
        # Create folders for the images
        if not os.path.isdir(MNIST_CF_FOLDER):
            os.makedirs(MNIST_CF_FOLDER)
        
        self.c_matrix.plot()
        if isAugmented == 'TRUE':
            plt.title("{} Augmented Confusion Matrix".format(self.name))
            plt.savefig("{}Augmented {}.png".format(MNIST_CF_FOLDER, self.name))
            img = Image("{}Augmented {}.png".format(MNIST_CF_FOLDER, self.name))
        else:
            plt.title("{} Confusion Matrix".format(self.name))
            plt.savefig("{}{}.png".format(MNIST_CF_FOLDER, self.name))
            img = Image("{}{}.png".format(MNIST_CF_FOLDER, self.name))
        
        img.height = IMAGE_HEIGHT
        img.width = IMAGE_WIDTH
        return img

    # Method that prints the classification report to a CSV file
    def classificationReportCSV(self, isAugmented):
        # Create folder for the classification report
        if not os.path.isdir(MNIST_CR_FOLDER):
            os.makedirs(MNIST_CR_FOLDER)
            
        dataframe = pd.DataFrame(self.report).transpose()
        
        if isAugmented:
            fileName = "Augmented {} classification_report.csv".format(self.name)
        else:
            fileName = "{} classification_report.csv".format(self.name)    
        
        dataframe.to_csv(os.path.join(MNIST_CR_FOLDER,fileName), index = True)
 
    # Method creates the augmented dataset
    def augmentData(self):
        X_train_augmented = [image for image in self.X_train]
        y_train_augmented = [image for image in self.y_train]
        
        for dx, dy in ((1,0), (-1,0), (0,1), (0,-1)):
            for image, label in zip(self.X_train, self.y_train):
                X_train_augmented.append(self.shiftImage(image, dx, dy))
                y_train_augmented.append(label)
                
        # Shuffle the dataset
        shuffle_idx = np.random.permutation(len(X_train_augmented))
        X_train_augmented = np.array(X_train_augmented)[shuffle_idx]
        y_train_augmented = np.array(y_train_augmented)[shuffle_idx]
        
        self.X_train = X_train_augmented
        self.y_train = y_train_augmented
     
    # Method that runs the classifier's information   
    def run(self, classifier_model, isAugmented):
        self.dateAndTime()
        self.model = classifier_model
        self.name = self.model.__class__.__name__ 
        
        if isAugmented:
            self.augmentData() # convert the xy variables to augmented
            self.calculateClassifierMetrics()
            self.printToSheet('TRUE')
            self.classificationReportCSV(True)
        else:
            self.calculateClassifierMetrics()
            self.printToSheet('FALSE')
            self.classificationReportCSV(False)